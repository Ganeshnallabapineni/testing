import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
book=openpyxl.load_workbook("/home/student/Documents/openpyxl.xlsx")
sheet=book.active
sheet.cell(row=1,column=2)
print(cell.value)
sheet.cell(row=2,column=2).value="rahul"
print(sheet.cell(row=2,column=2).value)
print(sheet.max_row)
print(sheet.max_column)
print(sheet['A5'].value)

for i in range(1,sheet.max_row+1):
	if sheet.cell(row=i,column=1).value=="t2":
		for j in range(1,sheet.max_coloumn+1):
			print(sheet.cell(row=i,column=j).value)


